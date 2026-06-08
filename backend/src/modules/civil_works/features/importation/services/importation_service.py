import os
import uuid
import openpyxl

from datetime import datetime, date
from typing import BinaryIO, Any, cast
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from anyio.to_thread import run_sync
from structlog import get_logger

from src.core.settings.constants import ROOT_PATH

from src.modules.civil_works.domain.models.excel_upload import ExcelUpload
from src.modules.civil_works.domain.models.civil_work import CivilWork
from src.modules.civil_works.features.importation.repositories.importation_repository import ImportationRepository



logger = get_logger(__name__)

MAX_EMPTY_STREAK = 10  # Detenerse tras 10 filas vacías o sin N° Trabajo


class ImportationService:
    def __init__(self, repo: ImportationRepository) -> None:
        self.repo = repo
        self.upload_dir = str(ROOT_PATH / "uploads")


    async def process_upload(self, user_id: int, filename: str, file_content: BinaryIO) -> None:
        # Guardar archivo físicamente para auditoría
        if not os.path.exists(self.upload_dir):
            os.makedirs(self.upload_dir, exist_ok=True)

        unique_filename: str = f"{uuid.uuid4()}_{filename}"
        file_path: str = os.path.join(self.upload_dir, unique_filename)

        file_content.seek(0)
        with open(file_path, "wb") as f:
            f.write(file_content.read())

        try:
            logger.debug("Starting proccess in a separated thread...")

            # Delegar el procesamiento pesado (CPU bound) a un hilo separado
            civil_works_data: list[dict[str, Any]] = await run_sync(
                self._sync_process_excel, file_path
            )


            logger.debug("Deleting previous uploads...")

            # Operación en BD (I/O asíncrono)
            await self.repo.delete_previous_uploads()


            logger.debug("Adding civil works to upload...")

            new_upload = ExcelUpload(
                id_user=user_id,
                filename=filename,
                file_path=file_path,
            )

            new_upload.civil_works = [CivilWork(**item) for item in civil_works_data]

            logger.debug("Saving upload to database...")
            await self.repo.save_upload(new_upload)
            await self.repo.db.commit()

            logger.debug("Upload saved!")

        except Exception as e:
            raise e


    def _sync_process_excel(self, file_path: str) -> list[dict[str, Any]]:
        """Procesamiento sincrónico de Excel para ejecutar en un hilo separado."""
        # Usar read_only=True para mejorar significativamente el rendimiento y memoria
        workbook: Workbook = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
        sheet: Worksheet = cast(Worksheet, workbook.active)

        if sheet is None:
            if not workbook.sheetnames:
                raise ValueError("El archivo Excel no contiene hojas.")
            sheet = cast(Worksheet, workbook[workbook.sheetnames[0]])

        header_row_idx = 1
        headers = []
        found_headers = False

        logger.debug("Loading excel headers...")

        for row_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            row_values: list[str] = [str(cell).strip().upper() if cell is not None else "" for cell in row]
            logger.debug(f"{row_idx=}")
            if "N° TRABAJO" in row_values or "N° DE TRABAJO" in row_values:
                headers: list[str] = row_values
                header_row_idx: int = row_idx
                found_headers = True
                break

        if not found_headers:
            logger.debug("Error.")
            raise ValueError("No se encontró la fila de cabeceras (buscando 'N° trabajo')")


        logger.debug("Load ended!")
        mapping: dict[str, int] = self._get_column_mapping(headers)


        logger.debug("Loading excel rows...")
        civil_works_data: list[dict[str, Any]] = []

        empty_streak = 0

        for row_idx, row in enumerate(sheet.iter_rows(min_row=header_row_idx + 1, values_only=True), start=header_row_idx + 1):
            # Si la fila es nula o falta el número de trabajo, aumentamos la racha

            if all(val is None for val in row) or row[mapping["numero_trabajo"]] is None:
                empty_streak += 1

                if empty_streak >= MAX_EMPTY_STREAK:
                    logger.debug("Empty streak reached, stopping", row_idx=row_idx)
                    break

                continue

            # Reiniciar racha si hay datos válidos
            empty_streak = 0

            if row_idx%100 == 0:
                logger.debug("Loading progress", row_idx=row_idx)

            data: dict[str, Any] = self._parse_row(row, mapping, row_idx)
            civil_works_data.append(data)

        logger.debug("Row loading ended!", row_count=row_idx)

        return civil_works_data


    def _get_column_mapping(self, headers: list[str]) -> dict[str, int]:
        """Busca los índices de las columnas basadas en nombres esperados."""
        mapping: dict[str, Any] = {}

        expected: dict[str, list[str]] = {
            "numero_trabajo": ["N° TRABAJO", "N° DE TRABAJO", "NUMERO TRABAJO", "Nº TRABAJO", "NRO TRABAJO"],
            "semana": ["SEMANA"],
            "fecha": ["FECHA"],
            "contratista": ["CONTRATISTA"],
            "detalle_trabajos": ["DETALLE TRABAJOS", "DETALLE"],
            "tipo_trabajo": ["TIPO TRABAJO", "TIPO DE TRABAJO"],
            "forma_pago": ["FORMA PAGO", "FORMA DE PAGO", "PAGO"],
            "agenda_tipo_levantamiento": ["AGENDA O TIPO LEVANTAMIENTO", "AGENDA", "TIPO LEVANTAMIENTO"],
            "zonal": ["ZONAL", "ZONA"],
            "localidad": ["LOCALIDAD", "COMUNA"],
            "region": ["REGIÓN", "REGION"],
            "edificio_instalacion": ["EDIFICIO/INSTALACIÓN", "EDIFICIO", "INSTALACION", "EDIFICIO/INSTALACION"],
            "estado_trabajos": ["ESTADO TRABAJOS", "ESTADO"],
            "factura": ["FACTURA"],
            "hh": ["HH"],
            "monto_neto": ["MONTO NETO", "VALOR NETO", "NETO"],
            "monto_hh": ["MONTO HH", "VALOR HH"],
        }


        for field, aliases in expected.items():
            found = False
            for alias in aliases:
                if alias in headers:
                    mapping[field] = headers.index(alias)
                    found = True
                    break

            # Campos obligatorios
            if not found and field in ["numero_trabajo", "fecha", "contratista", "tipo_trabajo", "zonal", "localidad", "region", "estado_trabajos", "monto_neto"]:
                 raise ValueError(f"Columna obligatoria no encontrada: {field} (buscado: {aliases})")

        return mapping


    def _parse_row(self, row: tuple[Any, ...], mapping: dict[str, int], row_idx: int) -> dict[str, Any]:
        """Convierte una fila de Excel en un diccionario para CivilWork."""
        try:
            return {
                "numero_trabajo": int(row[mapping["numero_trabajo"]]) if row[mapping["numero_trabajo"]] is not None else 0,
                "semana": int(row[mapping["semana"]]) if mapping.get("semana") is not None and row[mapping["semana"]] is not None else None,
                "fecha": self._to_date(row[mapping["fecha"]]),
                "contratista": str(row[mapping["contratista"]]),
                "detalle_trabajos": str(row[mapping["detalle_trabajos"]]) if mapping.get("detalle_trabajos") is not None and row[mapping["detalle_trabajos"]] is not None else None,
                "tipo_trabajo": str(row[mapping["tipo_trabajo"]]),
                "forma_pago": str(row[mapping.get("forma_pago", -1)]) if mapping.get("forma_pago") is not None and row[mapping["forma_pago"]] is not None else "N/A",
                "agenda_tipo_levantamiento": str(row[mapping["agenda_tipo_levantamiento"]]) if mapping.get("agenda_tipo_levantamiento") is not None and row[mapping["agenda_tipo_levantamiento"]] is not None else None,
                "zonal": str(row[mapping["zonal"]]),
                "localidad": str(row[mapping["localidad"]]),
                "region": str(row[mapping["region"]]),
                "edificio_instalacion": str(row[mapping["edificio_instalacion"]]) if mapping.get("edificio_instalacion") is not None and row[mapping["edificio_instalacion"]] is not None else None,
                "estado_trabajos": str(row[mapping["estado_trabajos"]]),
                "factura": str(row[mapping["factura"]]) if mapping.get("factura") is not None and row[mapping["factura"]] is not None else None,
                "hh": str(row[mapping["hh"]]) if mapping.get("hh") is not None and row[mapping["hh"]] is not None else None,
                "monto_neto": float(row[mapping["monto_neto"]]) if row[mapping["monto_neto"]] is not None else 0.0,
                "monto_hh": float(row[mapping["monto_hh"]]) if mapping.get("monto_hh") is not None and row[mapping["monto_hh"]] is not None else None,
            }

        except Exception as e:
            raise ValueError(f"Error en fila {row_idx}: {str(e)}")


    def _to_date(self, value: datetime | date | str) -> date:
        if isinstance(value, datetime):
            return value.date()

        if isinstance(value, date):
            return value

        if isinstance(value, str):
            value_clean: str = value.strip()

            try:
                return datetime.strptime(value_clean, "%d-%m-%Y").date()
            except ValueError as e:
                raise Exception(f"Formato de fecha inválido: {value_clean}")

        raise ValueError(f"Formato de fecha inválido: {value}")

